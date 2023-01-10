import mysql.connector
import pandas as pd
import os
import numpy as np
import openpyxl
import requests
from io import BytesIO

def get_unique_fields(folder):
    sources = []
    style = []
    hobby = []
    for filename in os.listdir(folder):
        if(filename.endswith('.xlsx')):
            df = pd.read_excel('dataset/'+filename, engine='openpyxl')
            if "Source" in df:
                sources = sources + df["Source"].values.tolist()
            if "Hobby" in df:
                hobby = hobby + df["Hobby"].values.tolist()
            if "Style 1" in df:
                style = style + df["Style 1"].values.tolist()

    return np.unique(sources),np.unique(hobby),np.unique(style)

sources,hobby,style = get_unique_fields('dataset')

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="enessefa55",
  port="8687",
)

mycursor = mydb.cursor(buffered=True)

mycursor.execute("CREATE DATABASE IF NOT EXISTS ProjectDb")

mycursor.execute("USE ProjectDb")

mycursor.execute("CREATE TABLE IF NOT EXISTS Source (Id INT AUTO_INCREMENT PRIMARY KEY, Name VARCHAR(255))")

mycursor.execute("CREATE TABLE IF NOT EXISTS Hobby (Id INT AUTO_INCREMENT PRIMARY KEY, Name VARCHAR(255))")

mycursor.execute("CREATE TABLE IF NOT EXISTS Style (Id INT AUTO_INCREMENT PRIMARY KEY, Name VARCHAR(255))")

query = "INSERT INTO Source (Name) VALUES (%s)"
mycursor.execute("Select Count(*) from Source")
temp = mycursor.fetchall()

if(temp[0][0] == 0):
    for row in sources:
        val = (str(row),)
        mycursor.execute(query,val)

    mydb.commit()

query = "INSERT INTO Hobby (Name) VALUES (%s)"
mycursor.execute("Select Count(*) from Hobby")
temp = mycursor.fetchall()
if(temp[0][0] == 0):
    for row in hobby:
        val = (str(row),)
        mycursor.execute(query,val)

    mydb.commit()

query = "INSERT INTO Style (Name) VALUES (%s)"
mycursor.execute("Select Count(*) from Style")
temp = mycursor.fetchall()
if(temp[0][0] == 0):
    for row in style:
        val = (str(row),)
        mycursor.execute(query,val)

    mydb.commit()


wallpaper_df = pd.read_excel('dataset/wallpaper.xlsx')
wallpaper_df = wallpaper_df[["Name","Source","Buy","Sell","Color 1","HHA Concept 1","Ceiling Type","Tag"]]

mycursor.execute("CREATE TABLE IF NOT EXISTS Wallpaper (Id INT AUTO_INCREMENT PRIMARY KEY, Name VARCHAR(255),Image LONGBLOB, SourceId INT, Buy int, Sell int, Color VARCHAR(255),Concept VARCHAR(255),Type VARCHAR(255),Tag VARCHAR(255), FOREIGN KEY (SourceId) REFERENCES Source(Id) ON DELETE SET NULL ON UPDATE CASCADE)")

wallpaper_df.loc[wallpaper_df["Buy"] == "NFS", "Buy"] = -1
wallpaper_df.loc[wallpaper_df["Sell"] == "NFS", "Sell"] = -1
query = "INSERT INTO Wallpaper (Name, Image, SourceId, Buy, Sell, Color, Concept, Type, Tag) VALUES (%s, %s, (SELECT Id from Source WHERE Name=%s),%s,%s,%s,%s,%s,%s)"
mycursor.execute("Select Count(*) from Wallpaper")
temp = mycursor.fetchall()
if(temp[0][0] == 0):
    images_wallpaper = []
    wb = openpyxl.load_workbook('dataset/wallpaper.xlsx')
    ws = wb['Sheet1']
    for col in ws['B'][1:]:
        url = col.value[8:-2]
        response = requests.get(url)
        images_wallpaper.append(BytesIO(response.content).getbuffer().tobytes())

    for row,img in zip(wallpaper_df.iterrows(),images_wallpaper):
        val = (row[1][0],img,row[1][1],int(row[1][2]),int(row[1][3]),row[1][4],row[1][5],row[1][6],row[1][7],)
        mycursor.execute(query,val)
    mydb.commit()


floor_df = pd.read_excel('dataset/floors.xlsx')
floor_df = floor_df[["Name","Source","Buy","Sell","Color 1","HHA Concept 1","Tag"]]

mycursor.execute("CREATE TABLE IF NOT EXISTS Floor (Id INT AUTO_INCREMENT PRIMARY KEY, Name VARCHAR(255),Image LONGBLOB, SourceId INT, Buy int, Sell int, Color VARCHAR(255),Concept VARCHAR(255),Tag VARCHAR(255),FOREIGN KEY (SourceId) REFERENCES Source(Id) ON DELETE SET NULL ON UPDATE CASCADE)")

floor_df.loc[floor_df["Buy"] == "NFS", "Buy"] = -1
floor_df.loc[floor_df["Sell"] == "NFS", "Sell"] = -1
query = "INSERT INTO Floor (Name, Image, SourceId, Buy, Sell, Color, Concept, Tag) VALUES (%s, %s, (SELECT Id from Source WHERE Name=%s),%s,%s,%s,%s,%s)"
mycursor.execute("Select Count(*) from Floor")
temp = mycursor.fetchall()
if(temp[0][0] == 0):
    images_floors = []
    wb = openpyxl.load_workbook('dataset/floors.xlsx')
    ws = wb['Sheet1']
    for col in ws['B'][1:]:
        url = col.value[8:-2]
        response = requests.get(url)
        images_floors.append(BytesIO(response.content).getbuffer().tobytes())
                
    for row,img in zip(floor_df.iterrows(),images_floors):
        val = (row[1][0],img,row[1][1],int(row[1][2]),int(row[1][3]),row[1][4],row[1][5],row[1][6],)
        mycursor.execute(query,val)
    mydb.commit()


print("Wallpaper and Floor tables created!")

villagers_df = pd.read_excel("dataset/villagers.xlsx")

villagers_df = villagers_df[["Name","Gender","Favorite Song","Personality","Style 1","Hobby"]]

mycursor.execute("CREATE TABLE IF NOT EXISTS Villagers (Id INT AUTO_INCREMENT PRIMARY KEY, Name VARCHAR(255),Image LONGBLOB, Gender VARCHAR(255), FavoriteSong VARCHAR(255), Personality VARCHAR(255),Style_id INT, Hobby_id INT, FOREIGN KEY (Style_id) REFERENCES Style(Id) ON DELETE SET NULL ON UPDATE CASCADE, FOREIGN KEY (Hobby_id) REFERENCES Hobby(Id) ON DELETE SET NULL ON UPDATE CASCADE)")

query = "INSERT INTO Villagers (Name,Image ,Gender,FavoriteSong, Personality, Style_id, Hobby_id) VALUES (%s,%s,%s,%s,%s,(SELECT id from Style WHERE Name=%s),(SELECT id from Hobby WHERE Name=%s))"
mycursor.execute("Select Count(*) from Villagers")
temp = mycursor.fetchall()
if(temp[0][0] == 0):
    images = []
    wb = openpyxl.load_workbook('dataset/villagers.xlsx')
    ws = wb['Sheet1']
    for col in ws['B'][1:]:
        url = col.value[8:-2]
        response = requests.get(url)
        images.append(BytesIO(response.content).getbuffer().tobytes())

    for row,image in zip(villagers_df.iterrows(),images) :
        val = (row[1][0],image,row[1][1],(row[1][2]),(row[1][3]),row[1][4],row[1][5],)
        mycursor.execute(query,val)
    mydb.commit()

tools_df = pd.read_excel("dataset/toolsgoods.xlsx")

tools_df = tools_df[["Name","Source","Buy","Sell","Variation"]]

mycursor.execute("CREATE TABLE IF NOT EXISTS Tools (Id INT AUTO_INCREMENT PRIMARY KEY,SourceId INT, Name VARCHAR(255),Image LONGBLOB, Buy int, Sell int, Color VARCHAR(255), FOREIGN KEY (SourceId) REFERENCES Source(Id) ON DELETE SET NULL ON UPDATE CASCADE)")



tools_df.loc[tools_df["Buy"] == "NFS", "Buy"] = -1
tools_df.loc[tools_df["Sell"] == "NFS", "Sell"] = -1
tools_df.loc[tools_df["Buy"] == "None", "Buy"] = -1
tools_df.loc[tools_df["Sell"] == "None", "Sell"] = -1

tools_df.fillna(-1,inplace=True)
query = "INSERT INTO Tools (Name,Image, SourceId, Buy, Sell, Color) VALUES (%s,%s,(SELECT id from Source WHERE Name=%s),%s,%s,%s)"
mycursor.execute("Select Count(*) from Tools")
temp = mycursor.fetchall()
print(temp[0][0])
if(temp[0][0] == 0):
    images_tools = []
    wb = openpyxl.load_workbook('dataset/toolsgoods.xlsx')
    ws = wb['Sheet1']
    for col in ws['B'][1:]:
        url = col.value[8:-2]
        response = requests.get(url)
        images_tools.append(BytesIO(response.content).getbuffer().tobytes())
    for row,image in zip(tools_df.iterrows(),images_tools):
        val = (row[1][0],image,row[1][1],int(row[1][2]),int(row[1][3]),row[1][4],)
        mycursor.execute(query,val)
    mydb.commit()


print("Villager and Tools tables are created!")
# busra tables

# accessories dataframe
accessoriesdf = pd.read_excel('dataset/accessories.xlsx')
accessoriesdf = accessoriesdf[["Name", "Source","Buy", "Sell", "Color 1", "Style 1", "Label Themes"]].drop_duplicates()
themes = []

# themes are selected from dataframe
for row in accessoriesdf.iterrows():
    temp = row[1][6].split("; ")
    
    themes += temp
themes = np.unique(themes)

# themes table creation if not exists
mycursor.execute("CREATE TABLE IF NOT EXISTS Themes (Id INT AUTO_INCREMENT PRIMARY KEY, NAME VARCHAR(255))")

# if table is empty fill it
mycursor.execute("Select Count(*) from Themes")
temp = mycursor.fetchall()
if(temp[0][0] == 0):
    # themes table creation
    query = "INSERT INTO Themes (Name) VALUES (%s)"
    for row in themes:
        val = (str(row),)
        mycursor.execute(query, val)

    mydb.commit()

# if not exists create accessories table
mycursor.execute("CREATE TABLE IF NOT EXISTS Accessories (Id INT AUTO_INCREMENT PRIMARY KEY, SourceId INT, Name VARCHAR(255), Image LONGBLOB, Buy int, Sell int, Color VARCHAR(255), Stylish VARCHAR(255), FOREIGN KEY (SourceId) REFERENCES Source(Id) ON DELETE SET NULL ON UPDATE CASCADE)")



accessoriesdf.loc[accessoriesdf["Buy"] == "NFS", "Buy"] = -1
accessoriesdf.loc[accessoriesdf["Sell"] == "NFS", "Sell"] = -1
query = "INSERT INTO Accessories (SourceId, Name, Image, Buy, Sell, Color, Stylish) VALUES ((SELECT id from Source WHERE Name=%s),%s,%s,%s,%s,%s,%s)"

# if accessories if empty, insert the rows
mycursor.execute("Select Count(*) from Accessories")
temp = mycursor.fetchall()
if(temp[0][0] == 0):
    images_accessories = []
    wb = openpyxl.load_workbook('dataset/accessories.xlsx')
    ws = wb['Sheet1']
    for col in ws['B'][1:]:
        url = col.value[8:-2]
        response = requests.get(url)
        images_accessories.append(BytesIO(response.content).getbuffer().tobytes())
    for row,img in zip(accessoriesdf.iterrows(),images_accessories):
        val = (row[1][1],row[1][0],img,int(row[1][2]),int(row[1][3]),row[1][4],row[1][5])
        mycursor.execute(query,val)
    mydb.commit()


# many to many table if not exists create accessories-themes
mycursor.execute("CREATE TABLE IF NOT EXISTS AccessoriesThemes (Id INT AUTO_INCREMENT PRIMARY KEY, AccessoriesId INT, ThemeId INT, FOREIGN KEY (AccessoriesId) REFERENCES Accessories(Id) ON DELETE SET NULL ON UPDATE CASCADE, FOREIGN KEY (ThemeId) REFERENCES Themes(Id) ON DELETE SET NULL ON UPDATE CASCADE)")


query = "INSERT INTO AccessoriesThemes (AccessoriesId, ThemeId) VALUES ((SELECT id from Accessories WHERE Name=%s AND Color=%s AND Stylish=%s), (SELECT id from Themes WHERE Name=%s))"

# if table is empty insert rows
mycursor.execute("Select Count(*) from AccessoriesThemes")
temp = mycursor.fetchall()
if(temp[0][0] == 0):
    for row in accessoriesdf.iterrows():
        name = row[1][0]
        color = row[1][4]
        style= row[1][5]
        themes = row[1][6].split("; ")

        for theme in themes:
            val = (name, color,style ,theme)
            mycursor.execute(query,val)

    mydb.commit()


# dress up dataframe
dressUp_df = pd.read_excel("dataset/dress-up.xlsx")
dressUp_df = dressUp_df[["Name", "Source","Buy", "Sell", "Color 1", "Style 1", "Primary Shape"]].drop_duplicates()

# if not exists create dress up table
mycursor.execute("CREATE TABLE IF NOT EXISTS DressUp (Id INT AUTO_INCREMENT PRIMARY KEY, SourceId INT, Name VARCHAR(255), Image LONGBLOB, Buy int, Sell int, Color VARCHAR(255), Stylish VARCHAR(255), Shape VARCHAR(255), FOREIGN KEY (SourceId) REFERENCES Source(Id) ON DELETE SET NULL ON UPDATE CASCADE)")

dressUp_df.loc[dressUp_df["Buy"] == "NFS", "Buy"] = -1
dressUp_df.loc[dressUp_df["Sell"] == "NFS", "Sell"] = -1
query = "INSERT INTO DressUp (SourceId, Name, Image, Buy, Sell, Color, Stylish, Shape) VALUES ((SELECT id from Source WHERE Name=%s),%s,%s,%s,%s,%s,%s, %s)"

# if table is empty insert rows
mycursor.execute("Select Count(*) from DressUp")
temp = mycursor.fetchall()
if(temp[0][0] == 0):
    images_dressup = []
    wb = openpyxl.load_workbook('dataset/dress-up.xlsx')
    ws = wb['Sheet1']
    for col in ws['B'][1:]:
        url = col.value[8:-2]
        response = requests.get(url)
        images_dressup.append(BytesIO(response.content).getbuffer().tobytes())
    for row,img in zip(dressUp_df.iterrows(),images_dressup):
        # float integer error (nan float to integer)
        # buy attribute
        try:
            buy = int(row[1][2])
        except:
            buy = 0
        # sell attribute
        try:
            sell = int(row[1][3])
        except:
            sell = 0
        val = (row[1][1],row[1][0],img,buy,sell,row[1][4],row[1][5],row[1][6])
        mycursor.execute(query,val)
    mydb.commit()

#print that tables are created
print("Accessories and DressUp tables are created!")


print("Database initialized succesfully")
mycursor.close()
mydb.close()